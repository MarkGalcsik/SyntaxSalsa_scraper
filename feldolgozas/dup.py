"""
highlight_duplicate_skus.py
----------------------------
Megkeresi a duplikált SKU-kat a 'Main' munkalap D oszlopában,
és kék háttérszínnel emeli ki őket.

Használat:
    python highlight_duplicate_skus.py <excel_fajl.xlsx>

Az eredmény ugyanabba a fájlba kerül vissza (felülírja).
Ha szeretnéd megtartani az eredetit, add meg a kimeneti fájlt is:
    python highlight_duplicate_skus.py input.xlsx output.xlsx
"""

import sys
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BLUE_FILL = PatternFill(fill_type="solid", start_color="ADD8E6", end_color="ADD8E6")  # világoskék
CLEAR_FILL = PatternFill(fill_type=None)  # kiemelés törlése

def process_column(ws, col_index: int):
    cells = [row[0] for row in ws.iter_rows(min_col=col_index, max_col=col_index)]
    values = [str(c.value).strip() for c in cells if c.value is not None and str(c.value).strip() != ""]
    duplicates = {sku for sku, cnt in Counter(values).items() if cnt > 1}

    highlighted = 0
    for cell in cells:
        val = str(cell.value).strip() if cell.value is not None else ""
        if val in duplicates:
            cell.fill = BLUE_FILL
            highlighted += 1
        else:
            cell.fill = CLEAR_FILL

    return highlighted, duplicates

def highlight_duplicate_skus(input_path: str, output_path: str | None = None):
    wb = load_workbook(input_path)

    if "Main" not in wb.sheetnames:
        raise ValueError(f"Nem található 'Main' nevű munkalap a fájlban: {input_path}")

    ws = wb["Main"]

    d_highlighted, d_dupes = process_column(ws, col_index=4)   # D oszlop
    l_highlighted, l_dupes = process_column(ws, col_index=12)  # L oszlop

    save_path = output_path or input_path
    wb.save(save_path)

    print(f"✅ D oszlop: {d_highlighted} cella kiemelve, {len(d_dupes)} egyedi duplikált SKU.")
    print(f"✅ L oszlop: {l_highlighted} cella kiemelve, {len(l_dupes)} egyedi duplikált SKU.")
    print(f"📁 Mentve: {save_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Használat: python highlight_duplicate_skus.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) >= 3 else None

    highlight_duplicate_skus(input_file, output_file)